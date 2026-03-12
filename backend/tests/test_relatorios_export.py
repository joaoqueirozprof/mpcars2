from datetime import datetime
from io import BytesIO

from openpyxl import load_workbook

from app.models import DespesaLoja, LancamentoFinanceiro


def test_financeiro_export_xlsx_handles_legacy_store_expenses(client, admin_headers, db_session):
    db_session.add(
        DespesaLoja(
            categoria="Operacao",
            descricao="Despesa antiga sem mes/ano",
            valor=350.75,
            mes=None,
            ano=None,
            data=datetime(2026, 3, 11, 10, 30),
        )
    )
    db_session.add(
        LancamentoFinanceiro(
            data=datetime(2026, 3, 12).date(),
            tipo="receita",
            categoria="Contrato",
            descricao="Recebimento manual",
            valor=890.50,
            status="pago",
        )
    )
    db_session.commit()

    response = client.get(
        "/api/v1/relatorios/exportar/financeiro?formato=xlsx&data_inicio=2026-03-01&data_fim=2026-03-31",
        headers=admin_headers,
    )

    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response.content[:2] == b"PK"

    workbook = load_workbook(filename=BytesIO(response.content))
    assert "Resumo" in workbook.sheetnames
    assert "Despesas Loja" in workbook.sheetnames
    assert "Lançamentos" in workbook.sheetnames

    resumo = workbook["Resumo"]
    lancamentos = workbook["Lançamentos"]
    despesas_loja = workbook["Despesas Loja"]

    assert resumo["A2"].value == "03/2026"
    assert lancamentos["D2"].value == "Recebimento manual"
    assert despesas_loja["B2"].value == "Despesa antiga sem mes/ano"


def test_financeiro_export_rejects_invalid_date_range(client, admin_headers):
    response = client.get(
        "/api/v1/relatorios/exportar/financeiro?formato=xlsx&data_inicio=2026-03-31&data_fim=2026-03-01",
        headers=admin_headers,
    )

    assert response.status_code == 400
    assert "data_inicio" in response.json()["detail"]
