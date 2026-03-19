"""
Serviço de exportação de dados para CSV e XLSX.
Suporta exportação de clientes, veículos, contratos e relatórios financeiros.
"""

from io import BytesIO
from datetime import datetime, date
from decimal import Decimal
from typing import Optional, List, Tuple
import re

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models import (
    Contrato,
    Cliente,
    Veiculo,
    DespesaVeiculo,
    DespesaLoja,
    Seguro,
    ParcelaSeguro,
    IpvaRegistro,
    Multa,
    LancamentoFinanceiro,
)


class ExportacaoService:
    """Serviço centralizado para exportação de dados em CSV e XLSX."""

    # Estilos XLSX
    HEADER_FILL = PatternFill(start_color="3B5998", end_color="3B5998", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

    @staticmethod
    def _normalize_sheet_name(value: str) -> str:
        """Mantém nomes de abas compatíveis com Excel/openpyxl."""
        sanitized = re.sub(r"[\\/*?:\[\]]", "-", (value or "").strip()) or "Relatorio"
        return sanitized[:31]

    @staticmethod
    def _coerce_to_date(value: Optional[object]) -> Optional[date]:
        """Converte strings, datetime e date para date de forma segura."""
        if value in (None, ""):
            return None
        if isinstance(value, date) and not isinstance(value, datetime):
            return value
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, str):
            for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
                try:
                    return datetime.strptime(value, fmt).date()
                except ValueError:
                    continue
        return None

    @staticmethod
    def _extract_month_bucket(value: Optional[object]) -> Optional[str]:
        normalized = ExportacaoService._coerce_to_date(value)
        if not normalized:
            return None
        return normalized.strftime("%m/%Y")

    @staticmethod
    def _resolve_despesa_loja_date(despesa: DespesaLoja) -> Optional[date]:
        """
        Resolve a data de referência da despesa da loja.

        Existem registros legados que usam apenas `data`, enquanto outros
        guardam `mes` e `ano`. O export precisa aceitar ambos.
        """
        mes = getattr(despesa, "mes", None)
        ano = getattr(despesa, "ano", None)

        if isinstance(mes, int) and isinstance(ano, int):
            try:
                return date(ano, mes, 1)
            except ValueError:
                pass

        data_referencia = ExportacaoService._coerce_to_date(getattr(despesa, "data", None))
        if data_referencia:
            return data_referencia.replace(day=1)

        return None

    @staticmethod
    def _safe_float(value: Optional[object]) -> float:
        if value in (None, ""):
            return 0.0
        if isinstance(value, Decimal):
            return float(value)
        try:
            return float(value)
        except (TypeError, ValueError):
            return 0.0

    @staticmethod
    def _format_date(value: Optional[date]) -> str:
        """Formata data para DD/MM/AAAA."""
        if not value:
            return ""
        if isinstance(value, str):
            return value
        return value.strftime("%d/%m/%Y") if value else ""

    @staticmethod
    def _format_currency(value: Optional[float]) -> str:
        """Formata valor monetário com R$ e 2 casas decimais."""
        if value is None or value == "":
            return "R$ 0,00"
        try:
            num = float(value)
            return f"R$ {num:,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")
        except (ValueError, TypeError):
            return "R$ 0,00"

    @staticmethod
    def _format_percentage(value: Optional[float]) -> str:
        """Formata valor percentual com 2 casas decimais."""
        if value is None or value == "":
            return "0,00%"
        try:
            num = float(value)
            return f"{num:.2f}%".replace(".", ",")
        except (ValueError, TypeError):
            return "0,00%"

    @staticmethod
    def _create_csv_content(headers: List[str], rows: List[List], separator: str = ";") -> bytes:
        """Cria conteúdo CSV com BOM UTF-8."""
        content = "\ufeff"  # BOM UTF-8
        content += separator.join(f'"{h}"' for h in headers) + "\n"

        for row in rows:
            line = []
            for cell in row:
                cell_str = str(cell) if cell is not None else ""
                # Escapa aspas dentro do valor
                cell_str = cell_str.replace('"', '""')
                line.append(f'"{cell_str}"')
            content += separator.join(line) + "\n"

        return content.encode("utf-8")

    @staticmethod
    def _create_xlsx_workbook(
        sheet_name: str,
        headers: List[str],
        rows: List[List],
    ) -> openpyxl.Workbook:
        """Cria workbook XLSX com formatação padrão."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = ExportacaoService._normalize_sheet_name(sheet_name)

        # Cabeçalhos
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = ExportacaoService.HEADER_FILL
            cell.font = ExportacaoService.HEADER_FONT
            cell.alignment = ExportacaoService.HEADER_ALIGNMENT

        # Dados
        for row_num, row in enumerate(rows, 2):
            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # Auto-width
        for col_num, header in enumerate(headers, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = max(
                len(str(header)) + 2, 12
            )

        return wb

    @staticmethod
    def _create_multi_sheet_xlsx(sheets_data: List[Tuple[str, List[str], List[List]]]) -> bytes:
        """Cria workbook XLSX com múltiplas abas."""
        wb = openpyxl.Workbook()
        if wb.sheetnames:
            wb.remove(wb.active)

        for sheet_name, headers, rows in sheets_data:
            ws = wb.create_sheet(ExportacaoService._normalize_sheet_name(sheet_name))

            # Cabeçalhos
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = ExportacaoService.HEADER_FILL
                cell.font = ExportacaoService.HEADER_FONT
                cell.alignment = ExportacaoService.HEADER_ALIGNMENT

            # Dados
            for row_num, row in enumerate(rows, 2):
                for col_num, value in enumerate(row, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.alignment = Alignment(horizontal="left", vertical="center")

            # Auto-width
            for col_num, header in enumerate(headers, 1):
                col_letter = openpyxl.utils.get_column_letter(col_num)
                max_value_length = max(
                    [len(str(header))]
                    + [len(str(row[col_num - 1])) for row in rows if len(row) >= col_num],
                    default=len(str(header)),
                )
                ws.column_dimensions[col_letter].width = min(max(max_value_length + 2, 12), 48)

        # Salva em BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def export_clientes(
        db: Session,
        formato: str = "xlsx",
        data_inicio: Optional[date] = None,
        data_fim: Optional[date] = None,
    ) -> BytesIO:
        """
        Exporta lista de clientes com informações de contratos.

        Args:
            db: Sessão do banco de dados
            formato: 'xlsx' ou 'csv'
            data_inicio: Filtro de data inicial de cadastro
            data_fim: Filtro de data final de cadastro

        Returns:
            BytesIO com dados do arquivo
        """
        query = db.query(Cliente)

        if data_inicio:
            query = query.filter(Cliente.data_cadastro >= data_inicio)
        if data_fim:
            query = query.filter(Cliente.data_cadastro <= data_fim)

        clientes = query.all()

        headers = [
            "ID",
            "Nome",
            "CPF/CNPJ",
            "RG",
            "Telefone",
            "E-mail",
            "CNH Número",
            "CNH Categoria",
            "CNH Validade",
            "Endereço",
            "Cidade",
            "Estado",
            "Score",
            "Data Cadastro",
            "Total Contratos",
            "Valor Total Locado",
        ]

        rows = []
        for cliente in clientes:
            # Contagem de contratos
            total_contratos = db.query(func.count(Contrato.id)).filter(
                Contrato.cliente_id == cliente.id
            ).scalar() or 0

            # Soma do valor total
            valor_total = db.query(func.sum(Contrato.valor_total)).filter(
                Contrato.cliente_id == cliente.id
            ).scalar() or 0

            row = [
                cliente.id,
                cliente.nome or "",
                cliente.cpf or "",
                cliente.rg or "",
                cliente.telefone or "",
                cliente.email or "",
                cliente.numero_cnh or "",
                cliente.categoria_cnh or "",
                ExportacaoService._format_date(cliente.validade_cnh),
                cliente.endereco_residencial or "",
                cliente.cidade_residencial or "",
                cliente.estado_residencial or "",
                cliente.score or "",
                ExportacaoService._format_date(cliente.data_cadastro),
                int(total_contratos),
                ExportacaoService._format_currency(valor_total),
            ]
            rows.append(row)

        if formato.lower() == "csv":
            content = ExportacaoService._create_csv_content(headers, rows)
            output = BytesIO(content)
        else:
            wb = ExportacaoService._create_xlsx_workbook("Clientes", headers, rows)
            output = BytesIO()
            wb.save(output)

        output.seek(0)
        return output

    @staticmethod
    def export_veiculos(
        db: Session,
        formato: str = "xlsx",
        status_filter: Optional[str] = None,
    ) -> BytesIO:
        """
        Exporta lista de veículos com informações de contratos e ROI.

        Args:
            db: Sessão do banco de dados
            formato: 'xlsx' ou 'csv'
            status_filter: Filtrar por status específico

        Returns:
            BytesIO com dados do arquivo
        """
        query = db.query(Veiculo)

        if status_filter:
            query = query.filter(Veiculo.status == status_filter)

        veiculos = query.all()

        headers = [
            "ID",
            "Placa",
            "Marca",
            "Modelo",
            "Ano",
            "Cor",
            "KM Atual",
            "Status",
            "Categoria",
            "Valor Diária",
            "Total Contratos",
            "Receita Total",
            "ROI (%)",
        ]

        rows = []
        for veiculo in veiculos:
            # Contagem de contratos
            total_contratos = db.query(func.count(Contrato.id)).filter(
                Contrato.veiculo_id == veiculo.id
            ).scalar() or 0

            # Receita total
            receita_total = db.query(func.sum(Contrato.valor_total)).filter(
                Contrato.veiculo_id == veiculo.id
            ).scalar() or 0

            # Cálculo ROI
            roi = 0
            valor_aquisicao = getattr(veiculo, 'valor_aquisicao', None) or 0
            if valor_aquisicao:
                try:
                    roi = (float(receita_total) / float(valor_aquisicao)) * 100
                except (ValueError, ZeroDivisionError):
                    roi = 0

            valor_diaria = veiculo.valor_diaria or 0
            if not valor_diaria and total_contratos > 0:
                avg_diaria = db.query(func.avg(Contrato.valor_diaria)).filter(
                    Contrato.veiculo_id == veiculo.id
                ).scalar()
                valor_diaria = avg_diaria or 0

            row = [
                veiculo.id,
                veiculo.placa or "",
                veiculo.marca or "",
                veiculo.modelo or "",
                veiculo.ano or "",
                veiculo.cor or "",
                veiculo.km_atual or 0,
                veiculo.status or "",
                getattr(veiculo, "categoria", "") or "",
                ExportacaoService._format_currency(valor_diaria),
                int(total_contratos),
                ExportacaoService._format_currency(receita_total),
                ExportacaoService._format_percentage(roi),
            ]
            rows.append(row)

        if formato.lower() == "csv":
            content = ExportacaoService._create_csv_content(headers, rows)
            output = BytesIO(content)
        else:
            wb = ExportacaoService._create_xlsx_workbook("Veículos", headers, rows)
            output = BytesIO()
            wb.save(output)

        output.seek(0)
        return output

    @staticmethod
    def export_contratos(
        db: Session,
        formato: str = "xlsx",
        data_inicio: Optional[date] = None,
        data_fim: Optional[date] = None,
        status_filter: Optional[str] = None,
    ) -> BytesIO:
        """
        Exporta lista de contratos com detalhes de saída, devolução e custos.

        Args:
            db: Sessão do banco de dados
            formato: 'xlsx' ou 'csv'
            data_inicio: Filtro data inicial
            data_fim: Filtro data final
            status_filter: Filtrar por status

        Returns:
            BytesIO com dados do arquivo
        """
        query = db.query(Contrato).join(Cliente).join(Veiculo)

        if data_inicio:
            query = query.filter(Contrato.data_inicio >= data_inicio)
        if data_fim:
            query = query.filter(Contrato.data_inicio <= data_fim)
        if status_filter:
            query = query.filter(Contrato.status == status_filter)

        contratos = query.all()

        headers = [
            "ID",
            "Cliente",
            "CPF/CNPJ",
            "Veículo",
            "Data Saída",
            "Data Devolução Prevista",
            "Data Devolução Real",
            "KM Saída",
            "KM Retorno",
            "KM Percorrida",
            "Período Utilizado/Faturado",
            "Valor Diária",
            "Combustível Saída",
            "Combustível Retorno",
            "Avarias (R$)",
            "Desconto (R$)",
            "Valor Total (R$)",
            "Status",
            "Tipo",
        ]

        rows = []
        for contrato in contratos:
            km_percorrida = 0
            if contrato.km_inicial is not None and contrato.km_final is not None:
                km_percorrida = contrato.km_final - contrato.km_inicial

            # Cálculo de diárias
            diarias = contrato.qtd_diarias or 0
            if not diarias and contrato.data_inicio and contrato.data_fim:
                delta = contrato.data_fim - contrato.data_inicio
                diarias = delta.days or 1

            cliente_nome = contrato.cliente.nome if contrato.cliente else ""
            cliente_cpf = contrato.cliente.cpf if contrato.cliente else ""
            veiculo_info = ""
            if contrato.veiculo:
                veiculo_info = f"{contrato.veiculo.placa or ''} {contrato.veiculo.modelo or ''}".strip()

            row = [
                contrato.id,
                cliente_nome,
                cliente_cpf,
                veiculo_info,
                ExportacaoService._format_date(contrato.data_inicio),
                ExportacaoService._format_date(contrato.data_fim),
                ExportacaoService._format_date(contrato.data_fim),
                contrato.km_inicial or 0,
                contrato.km_final or 0,
                km_percorrida,
                diarias,
                ExportacaoService._format_currency(contrato.valor_diaria or 0),
                ExportacaoService._format_currency(contrato.combustivel_saida or 0),
                ExportacaoService._format_currency(contrato.combustivel_retorno or 0),
                ExportacaoService._format_currency(contrato.valor_avarias or 0),
                ExportacaoService._format_currency(contrato.desconto or 0),
                ExportacaoService._format_currency(contrato.valor_total or 0),
                contrato.status or "",
                contrato.tipo or "",
            ]
            rows.append(row)

        if formato.lower() == "csv":
            content = ExportacaoService._create_csv_content(headers, rows)
            output = BytesIO(content)
        else:
            wb = ExportacaoService._create_xlsx_workbook("Contratos", headers, rows)
            output = BytesIO()
            wb.save(output)

        output.seek(0)
        return output

    @staticmethod
    def export_financeiro(
        db: Session,
        formato: str = "xlsx",
        data_inicio: Optional[date] = None,
        data_fim: Optional[date] = None,
    ) -> BytesIO:
        """Exporta relatorio financeiro completo com TODAS as fontes de dados."""
        from app.models import (
            Contrato, Cliente, Veiculo, DespesaContrato, DespesaVeiculo,
            DespesaLoja, Manutencao, Seguro, IpvaRegistro, Multa,
            LancamentoFinanceiro, Empresa, RelatorioNF,
        )
        from datetime import datetime as dt

        # ── Collect ALL records (same logic as PDF report) ──
        all_records = []

        # 1. Contratos PF
        for c in db.query(Contrato).all():
            if str(c.tipo or "").lower() == "empresa":
                continue
            cliente = db.query(Cliente).filter(Cliente.id == c.cliente_id).first()
            veiculo = db.query(Veiculo).filter(Veiculo.id == c.veiculo_id).first()
            all_records.append({
                "data": c.data_criacao.date() if c.data_criacao else None,
                "tipo": "Receita",
                "categoria": "Locacao",
                "descricao": "Contrato #{} - {}".format(c.numero, cliente.nome if cliente else "N/A"),
                "veiculo": "{} {}".format(veiculo.placa, veiculo.modelo) if veiculo else "",
                "valor": float(c.valor_total or 0),
                "status": "Pago" if c.status == "finalizado" else "Pendente",
            })

        # 2. NF empresa
        for nf in db.query(RelatorioNF).all():
            veiculo = db.query(Veiculo).filter(Veiculo.id == nf.veiculo_id).first()
            empresa = db.query(Empresa).filter(Empresa.id == nf.empresa_id).first() if nf.empresa_id else None
            periodo = ""
            if nf.periodo_inicio:
                periodo = nf.periodo_inicio.strftime("%d/%m/%y")
                if nf.periodo_fim:
                    periodo += " a " + nf.periodo_fim.strftime("%d/%m/%y")
            all_records.append({
                "data": nf.data_criacao.date() if nf.data_criacao else None,
                "tipo": "Receita",
                "categoria": "Faturamento Empresa",
                "descricao": "NF {} {} | {}".format(
                    veiculo.placa if veiculo else "", periodo, empresa.nome if empresa else ""
                ).strip(),
                "veiculo": "{} {}".format(veiculo.placa, veiculo.modelo) if veiculo else "",
                "valor": float(nf.valor_total_periodo or 0),
                "status": "Pago" if nf.pago else "Pendente",
            })

        # 3. Despesas Contrato
        for d in db.query(DespesaContrato).all():
            all_records.append({
                "data": d.data_registro.date() if d.data_registro else None,
                "tipo": "Despesa",
                "categoria": "Desp. Contrato",
                "descricao": d.descricao or "",
                "veiculo": "",
                "valor": float(d.valor or 0),
                "status": "Pago",
            })

        # 4. Despesas Veiculo
        for d in db.query(DespesaVeiculo).all():
            veiculo = db.query(Veiculo).filter(Veiculo.id == d.veiculo_id).first() if d.veiculo_id else None
            all_records.append({
                "data": d.data if d.data else None,
                "tipo": "Despesa",
                "categoria": "Desp. Veiculo",
                "descricao": d.descricao or "",
                "veiculo": "{} {}".format(veiculo.placa, veiculo.modelo) if veiculo else "",
                "valor": float(d.valor or 0),
                "status": "Pago",
            })

        # 5. Despesas Loja
        for d in db.query(DespesaLoja).all():
            all_records.append({
                "data": d.data if d.data else None,
                "tipo": "Despesa",
                "categoria": "Desp. Loja",
                "descricao": "{} - {}".format(d.categoria or "", d.descricao or "").strip(" -"),
                "veiculo": "",
                "valor": float(d.valor or 0),
                "status": "Pago",
            })

        # 6. Manutencoes
        for m in db.query(Manutencao).all():
            veiculo = db.query(Veiculo).filter(Veiculo.id == m.veiculo_id).first() if m.veiculo_id else None
            data_ref = m.data_realizada or (m.updated_at.date() if m.updated_at else None) or (m.data_criacao.date() if m.data_criacao else None)
            all_records.append({
                "data": data_ref,
                "tipo": "Despesa",
                "categoria": "Manutencao",
                "descricao": m.descricao or "",
                "veiculo": "{} {}".format(veiculo.placa, veiculo.modelo) if veiculo else "",
                "valor": float(m.custo or 0),
                "status": "Pago" if m.status == "concluida" else "Pendente",
            })

        # 7. Seguros
        for s in db.query(Seguro).all():
            veiculo = db.query(Veiculo).filter(Veiculo.id == s.veiculo_id).first() if s.veiculo_id else None
            all_records.append({
                "data": s.data_inicio if s.data_inicio else None,
                "tipo": "Despesa",
                "categoria": "Seguro",
                "descricao": s.seguradora or "Seguro",
                "veiculo": "{} {}".format(veiculo.placa, veiculo.modelo) if veiculo else "",
                "valor": float(s.valor or 0),
                "status": "Pago" if s.status == "ativo" else "Pendente",
            })

        # 8. IPVA
        for ip in db.query(IpvaRegistro).all():
            veiculo = db.query(Veiculo).filter(Veiculo.id == ip.veiculo_id).first() if ip.veiculo_id else None
            data_ref = ip.data_pagamento or ip.data_vencimento
            all_records.append({
                "data": data_ref,
                "tipo": "Despesa",
                "categoria": "IPVA",
                "descricao": "IPVA {}".format(ip.ano_referencia or ""),
                "veiculo": "{} {}".format(veiculo.placa, veiculo.modelo) if veiculo else "",
                "valor": float(ip.valor_ipva or ip.valor_pago or 0),
                "status": "Pago" if ip.status == "pago" else "Pendente",
            })

        # 9. Multas
        for ml in db.query(Multa).all():
            veiculo = db.query(Veiculo).filter(Veiculo.id == ml.veiculo_id).first() if ml.veiculo_id else None
            data_ref = ml.data_pagamento or ml.data_infracao
            all_records.append({
                "data": data_ref,
                "tipo": "Despesa",
                "categoria": "Multa",
                "descricao": ml.descricao or "Multa",
                "veiculo": "{} {}".format(veiculo.placa, veiculo.modelo) if veiculo else "",
                "valor": float(ml.valor or 0),
                "status": "Pago" if ml.status == "pago" else "Pendente",
            })

        # 10. Lancamentos manuais
        for lf in db.query(LancamentoFinanceiro).all():
            all_records.append({
                "data": lf.data if lf.data else None,
                "tipo": (lf.tipo or "despesa").capitalize(),
                "categoria": lf.categoria or "Outros",
                "descricao": lf.descricao or "",
                "veiculo": "",
                "valor": float(lf.valor or 0),
                "status": (lf.status or "pendente").capitalize(),
            })

        # ── Filter by date ──
        if data_inicio:
            all_records = [r for r in all_records if not r["data"] or r["data"] >= data_inicio]
        if data_fim:
            all_records = [r for r in all_records if not r["data"] or r["data"] <= data_fim]

        all_records.sort(key=lambda x: str(x.get("data") or ""), reverse=True)

        receitas = [r for r in all_records if r["tipo"] == "Receita"]
        despesas = [r for r in all_records if r["tipo"] == "Despesa"]
        total_receita = sum(r["valor"] for r in receitas)
        total_despesa = sum(r["valor"] for r in despesas)

        # Group despesas by category
        desp_por_cat = {}
        for r in despesas:
            cat = r.get("categoria") or "Outros"
            desp_por_cat[cat] = desp_por_cat.get(cat, 0) + r["valor"]

        if formato.lower() == "csv":
            return ExportacaoService._export_financeiro_csv(all_records, total_receita, total_despesa)
        else:
            return ExportacaoService._export_financeiro_xlsx(
                all_records, receitas, despesas,
                total_receita, total_despesa, desp_por_cat,
            )

    @staticmethod
    def _export_financeiro_csv(records, total_receita, total_despesa):
        """Export all financial records as CSV."""
        content = "\ufeff"  # BOM
        content += "RELATORIO FINANCEIRO COMPLETO\n"
        content += "Total Receitas;R$ {:,.2f}\n".format(total_receita)
        content += "Total Despesas;R$ {:,.2f}\n".format(total_despesa)
        content += "Lucro;R$ {:,.2f}\n\n".format(total_receita - total_despesa)

        headers = ["Data", "Tipo", "Categoria", "Descricao", "Veiculo", "Valor", "Status"]
        content += ";".join('"' + h + '"' for h in headers) + "\n"
        for r in records:
            data_str = r["data"].strftime("%d/%m/%Y") if r["data"] else ""
            line = [
                data_str,
                r["tipo"],
                r["categoria"],
                r["descricao"].replace('"', '""'),
                r["veiculo"],
                "{:.2f}".format(r["valor"]).replace(".", ","),
                r["status"],
            ]
            content += ";".join('"' + str(c) + '"' for c in line) + "\n"

        content += '\n"";"";"";"";"TOTAL RECEITAS";"{}";""\n'.format(
            "{:.2f}".format(total_receita).replace(".", ","))
        content += '"";"";"";"";"TOTAL DESPESAS";"{}";""\n'.format(
            "{:.2f}".format(total_despesa).replace(".", ","))
        content += '"";"";"";"";"LUCRO";"{}";""\n'.format(
            "{:.2f}".format(total_receita - total_despesa).replace(".", ","))

        output = BytesIO(content.encode("utf-8"))
        output.seek(0)
        return output

    @staticmethod
    def _export_financeiro_xlsx(records, receitas, despesas, total_receita, total_despesa, desp_por_cat):
        """Export formatted multi-tab Excel."""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        if wb.sheetnames:
            wb.remove(wb.active)

        # Style definitions
        thin_border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        green_font = Font(name="Calibri", bold=True, color="059669", size=11)
        red_font = Font(name="Calibri", bold=True, color="DC2626", size=11)
        total_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        total_font = Font(name="Calibri", bold=True, color="1E293B", size=12)
        zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        normal_font = Font(name="Calibri", size=10)
        curr_format = '#,##0.00'

        def write_sheet(ws, title, headers, rows, col_widths, value_col_idx=None, tipo_col_idx=None):
            # Title row
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            title_cell = ws.cell(row=1, column=1)
            title_cell.value = title
            title_cell.font = Font(name="Calibri", bold=True, size=14, color="1E3A5F")
            title_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[1].height = 30

            # Headers at row 2
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align
                cell.border = thin_border
            ws.row_dimensions[2].height = 22

            # Data rows
            for row_idx, row in enumerate(rows, 3):
                is_receita = False
                if tipo_col_idx is not None and len(row) > tipo_col_idx:
                    is_receita = str(row[tipo_col_idx]).lower() in ("receita",)

                for col_num, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_num)
                    cell.value = value
                    cell.font = normal_font
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

                    # Currency formatting
                    if value_col_idx is not None and col_num == value_col_idx + 1:
                        cell.number_format = curr_format
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        if is_receita:
                            cell.font = green_font
                        else:
                            cell.font = red_font

                    # Date formatting
                    if col_num == 1 and isinstance(value, date):
                        cell.number_format = "DD/MM/YYYY"
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                # Zebra striping
                if row_idx % 2 == 1:
                    for col_num in range(1, len(headers) + 1):
                        ws.cell(row=row_idx, column=col_num).fill = zebra_fill

            # Column widths
            for col_num, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(col_num)].width = width

            # Freeze header
            ws.freeze_panes = "A3"

            return len(rows) + 2  # last data row

        # ── ABA 1: RESUMO ──
        ws_resumo = wb.create_sheet("Resumo")
        ws_resumo.merge_cells("A1:B1")
        ws_resumo.cell(row=1, column=1).value = "RESUMO FINANCEIRO"
        ws_resumo.cell(row=1, column=1).font = Font(name="Calibri", bold=True, size=16, color="1E3A5F")
        ws_resumo.row_dimensions[1].height = 35

        summary_items = [
            ("Total de Receitas", total_receita, green_font, green_fill),
            ("Total de Despesas", total_despesa, red_font, red_fill),
            ("Lucro Liquido", total_receita - total_despesa,
             green_font if total_receita >= total_despesa else red_font,
             green_fill if total_receita >= total_despesa else red_fill),
            ("Margem de Lucro", (total_receita - total_despesa) / total_receita * 100 if total_receita > 0 else 0, None, None),
            ("Total de Registros", len(records), None, None),
        ]
        for i, (label, value, vfont, vfill) in enumerate(summary_items, 3):
            lc = ws_resumo.cell(row=i, column=1)
            lc.value = label
            lc.font = Font(name="Calibri", bold=True, size=12, color="475569")
            lc.border = thin_border
            lc.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

            vc = ws_resumo.cell(row=i, column=2)
            vc.value = value
            vc.border = thin_border
            if isinstance(value, float) and label != "Margem de Lucro":
                vc.number_format = 'R$ #,##0.00'
                vc.font = vfont or total_font
                if vfill:
                    vc.fill = vfill
            elif label == "Margem de Lucro":
                vc.number_format = '0.0"%"'
                vc.font = Font(name="Calibri", bold=True, size=12)
            else:
                vc.font = total_font

        # Despesas por categoria
        row_start = len(summary_items) + 5
        ws_resumo.cell(row=row_start, column=1).value = "DESPESAS POR CATEGORIA"
        ws_resumo.cell(row=row_start, column=1).font = Font(name="Calibri", bold=True, size=13, color="DC2626")
        for col in range(1, 4):
            cell = ws_resumo.cell(row=row_start + 1, column=col)
            cell.value = ["Categoria", "Valor", "% do Total"][col-1]
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = header_align
        for i, (cat, val) in enumerate(sorted(desp_por_cat.items(), key=lambda x: -x[1])):
            r = row_start + 2 + i
            ws_resumo.cell(row=r, column=1).value = cat
            ws_resumo.cell(row=r, column=1).font = normal_font
            ws_resumo.cell(row=r, column=1).border = thin_border
            vc = ws_resumo.cell(row=r, column=2)
            vc.value = val
            vc.number_format = 'R$ #,##0.00'
            vc.font = red_font
            vc.border = thin_border
            pc = ws_resumo.cell(row=r, column=3)
            pc.value = (val / total_despesa * 100) if total_despesa > 0 else 0
            pc.number_format = '0.0"%"'
            pc.font = normal_font
            pc.border = thin_border

        ws_resumo.column_dimensions["A"].width = 30
        ws_resumo.column_dimensions["B"].width = 22
        ws_resumo.column_dimensions["C"].width = 15

        # ── ABA 2: TODOS OS REGISTROS ──
        ws_all = wb.create_sheet("Todos os Registros")
        all_headers = ["Data", "Tipo", "Categoria", "Descricao", "Veiculo", "Valor", "Status"]
        all_rows = []
        for r in records:
            all_rows.append([
                r["data"],
                r["tipo"],
                r["categoria"],
                r["descricao"],
                r["veiculo"],
                r["valor"],
                r["status"],
            ])
        last_row = write_sheet(ws_all, "TODOS OS REGISTROS ({})".format(len(records)),
                               all_headers, all_rows, [14, 12, 18, 45, 20, 16, 12],
                               value_col_idx=5, tipo_col_idx=1)
        # Totals
        tr = last_row + 1
        for col in range(1, 8):
            ws_all.cell(row=tr, column=col).fill = total_fill
            ws_all.cell(row=tr, column=col).border = thin_border
        ws_all.cell(row=tr, column=5).value = "TOTAL RECEITAS"
        ws_all.cell(row=tr, column=5).font = total_font
        ws_all.cell(row=tr, column=6).value = total_receita
        ws_all.cell(row=tr, column=6).number_format = 'R$ #,##0.00'
        ws_all.cell(row=tr, column=6).font = green_font
        tr2 = tr + 1
        for col in range(1, 8):
            ws_all.cell(row=tr2, column=col).fill = total_fill
            ws_all.cell(row=tr2, column=col).border = thin_border
        ws_all.cell(row=tr2, column=5).value = "TOTAL DESPESAS"
        ws_all.cell(row=tr2, column=5).font = total_font
        ws_all.cell(row=tr2, column=6).value = total_despesa
        ws_all.cell(row=tr2, column=6).number_format = 'R$ #,##0.00'
        ws_all.cell(row=tr2, column=6).font = red_font
        tr3 = tr2 + 1
        for col in range(1, 8):
            ws_all.cell(row=tr3, column=col).fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
            ws_all.cell(row=tr3, column=col).border = thin_border
        ws_all.cell(row=tr3, column=5).value = "LUCRO"
        ws_all.cell(row=tr3, column=5).font = Font(name="Calibri", bold=True, size=13, color="1E3A5F")
        ws_all.cell(row=tr3, column=6).value = total_receita - total_despesa
        ws_all.cell(row=tr3, column=6).number_format = 'R$ #,##0.00'
        ws_all.cell(row=tr3, column=6).font = Font(name="Calibri", bold=True, size=13,
                                                     color="059669" if total_receita >= total_despesa else "DC2626")

        # ── ABA 3: RECEITAS ──
        if receitas:
            ws_rec = wb.create_sheet("Receitas")
            rec_headers = ["Data", "Categoria", "Descricao", "Veiculo", "Valor", "Status"]
            rec_rows = [[r["data"], r["categoria"], r["descricao"], r["veiculo"], r["valor"], r["status"]] for r in receitas]
            lr = write_sheet(ws_rec, "RECEITAS ({})".format(len(receitas)),
                             rec_headers, rec_rows, [14, 20, 45, 20, 16, 12], value_col_idx=4)
            tr = lr + 1
            for col in range(1, 7):
                ws_rec.cell(row=tr, column=col).fill = green_fill
                ws_rec.cell(row=tr, column=col).border = thin_border
            ws_rec.cell(row=tr, column=4).value = "TOTAL"
            ws_rec.cell(row=tr, column=4).font = total_font
            ws_rec.cell(row=tr, column=5).value = total_receita
            ws_rec.cell(row=tr, column=5).number_format = 'R$ #,##0.00'
            ws_rec.cell(row=tr, column=5).font = green_font

        # ── ABA 4: DESPESAS ──
        if despesas:
            ws_desp = wb.create_sheet("Despesas")
            desp_headers = ["Data", "Categoria", "Descricao", "Veiculo", "Valor", "Status"]
            desp_rows = [[r["data"], r["categoria"], r["descricao"], r["veiculo"], r["valor"], r["status"]] for r in despesas]
            lr = write_sheet(ws_desp, "DESPESAS ({})".format(len(despesas)),
                             desp_headers, desp_rows, [14, 20, 45, 20, 16, 12], value_col_idx=4)
            tr = lr + 1
            for col in range(1, 7):
                ws_desp.cell(row=tr, column=col).fill = red_fill
                ws_desp.cell(row=tr, column=col).border = thin_border
            ws_desp.cell(row=tr, column=4).value = "TOTAL"
            ws_desp.cell(row=tr, column=4).font = total_font
            ws_desp.cell(row=tr, column=5).value = total_despesa
            ws_desp.cell(row=tr, column=5).number_format = 'R$ #,##0.00'
            ws_desp.cell(row=tr, column=5).font = red_font

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output


    @staticmethod
    def _generate_resumo_mensal(
        db: Session,
        data_inicio: Optional[date],
        data_fim: Optional[date],
    ) -> List[List]:
        """Gera dados do resumo mensal."""
        rows = []

        # Busca contratos finalizados
        query = db.query(Contrato).filter(Contrato.status == "finalizado")

        if data_inicio:
            query = query.filter(Contrato.data_fim >= data_inicio)
        if data_fim:
            query = query.filter(Contrato.data_fim <= data_fim)

        contratos = query.all()

        # Agrupa por mês
        meses = {}
        for contrato in contratos:
            if contrato.data_fim:
                chave_mes = contrato.data_fim.strftime("%m/%Y")
                if chave_mes not in meses:
                    meses[chave_mes] = {"receita": 0, "despesa": 0}
                meses[chave_mes]["receita"] += float(contrato.valor_total or 0)

        # Despesas por mês
        despesas_veiculo = db.query(DespesaVeiculo)
        if data_inicio:
            despesas_veiculo = despesas_veiculo.filter(DespesaVeiculo.data >= data_inicio)
        if data_fim:
            despesas_veiculo = despesas_veiculo.filter(DespesaVeiculo.data <= data_fim)

        for despesa in despesas_veiculo.all():
            if despesa.data:
                chave_mes = despesa.data.strftime("%m/%Y")
                if chave_mes not in meses:
                    meses[chave_mes] = {"receita": 0, "despesa": 0}
                meses[chave_mes]["despesa"] += float(despesa.valor or 0)

        despesas_loja = db.query(DespesaLoja)
        for despesa in despesas_loja.all():
            referencia = ExportacaoService._resolve_despesa_loja_date(despesa)
            if not referencia:
                continue
            if data_inicio and referencia < data_inicio:
                continue
            if data_fim and referencia > data_fim:
                continue

            chave_mes = ExportacaoService._extract_month_bucket(referencia)
            if not chave_mes:
                continue
            if chave_mes not in meses:
                meses[chave_mes] = {"receita": 0, "despesa": 0}
            meses[chave_mes]["despesa"] += ExportacaoService._safe_float(despesa.valor)

        lancamentos = db.query(LancamentoFinanceiro)
        if data_inicio:
            lancamentos = lancamentos.filter(LancamentoFinanceiro.data >= data_inicio)
        if data_fim:
            lancamentos = lancamentos.filter(LancamentoFinanceiro.data <= data_fim)

        for lancamento in lancamentos.all():
            chave_mes = ExportacaoService._extract_month_bucket(lancamento.data)
            if not chave_mes:
                continue
            if chave_mes not in meses:
                meses[chave_mes] = {"receita": 0, "despesa": 0}

            tipo = (lancamento.tipo or "").lower()
            if tipo == "receita":
                meses[chave_mes]["receita"] += ExportacaoService._safe_float(lancamento.valor)
            else:
                meses[chave_mes]["despesa"] += ExportacaoService._safe_float(lancamento.valor)

        # Ordena por mês
        for chave_mes in sorted(meses.keys()):
            receita = meses[chave_mes]["receita"]
            despesa = meses[chave_mes]["despesa"]
            lucro = receita - despesa
            margem = (lucro / receita * 100) if receita > 0 else 0

            row = [
                chave_mes,
                ExportacaoService._format_currency(receita),
                ExportacaoService._format_currency(despesa),
                ExportacaoService._format_currency(lucro),
                ExportacaoService._format_percentage(margem),
            ]
            rows.append(row)

        return rows

    @staticmethod
    def _generate_receitas(
        db: Session,
        data_inicio: Optional[date],
        data_fim: Optional[date],
    ) -> List[List]:
        """Gera dados de receitas (contratos finalizados)."""
        query = db.query(Contrato).filter(Contrato.status == "finalizado")

        if data_inicio:
            query = query.filter(Contrato.data_fim >= data_inicio)
        if data_fim:
            query = query.filter(Contrato.data_fim <= data_fim)

        contratos = query.all()

        rows = []
        for contrato in contratos:
            km_percorrida = 0
            if contrato.km_inicial is not None and contrato.km_final is not None:
                km_percorrida = contrato.km_final - contrato.km_inicial

            diarias = contrato.qtd_diarias or 0
            if not diarias and contrato.data_inicio and contrato.data_fim:
                delta = contrato.data_fim - contrato.data_inicio
                diarias = delta.days or 1

            cliente_nome = contrato.cliente.nome if contrato.cliente else ""
            cliente_cpf = contrato.cliente.cpf if contrato.cliente else ""
            veiculo_info = ""
            if contrato.veiculo:
                veiculo_info = f"{contrato.veiculo.placa or ''} {contrato.veiculo.modelo or ''}".strip()

            row = [
                contrato.id,
                cliente_nome,
                cliente_cpf,
                veiculo_info,
                ExportacaoService._format_date(contrato.data_inicio),
                ExportacaoService._format_date(contrato.data_fim),
                ExportacaoService._format_date(contrato.data_fim),
                contrato.km_inicial or 0,
                contrato.km_final or 0,
                km_percorrida,
                diarias,
                ExportacaoService._format_currency(contrato.valor_diaria or 0),
                ExportacaoService._format_currency(contrato.combustivel_saida or 0),
                ExportacaoService._format_currency(contrato.combustivel_retorno or 0),
                ExportacaoService._format_currency(contrato.valor_avarias or 0),
                ExportacaoService._format_currency(contrato.desconto or 0),
                ExportacaoService._format_currency(contrato.valor_total or 0),
                contrato.status or "",
                contrato.tipo or "",
            ]
            rows.append(row)

        return rows

    @staticmethod
    def _generate_despesas_veiculos(
        db: Session,
        data_inicio: Optional[date],
        data_fim: Optional[date],
    ) -> List[List]:
        """Gera dados de despesas com veículos."""
        query = db.query(DespesaVeiculo).join(Veiculo)

        if data_inicio:
            query = query.filter(DespesaVeiculo.data >= data_inicio)
        if data_fim:
            query = query.filter(DespesaVeiculo.data <= data_fim)

        despesas = query.all()

        rows = []
        for despesa in despesas:
            veiculo_info = ""
            if despesa.veiculo:
                veiculo_info = f"{despesa.veiculo.placa or ''} {despesa.veiculo.modelo or ''}".strip()

            row = [
                veiculo_info,
                despesa.tipo or "",
                despesa.descricao or "",
                ExportacaoService._format_date(despesa.data),
                ExportacaoService._format_currency(despesa.valor or 0),
            ]
            rows.append(row)

        return rows

    @staticmethod
    def _generate_despesas_loja(
        db: Session,
        data_inicio: Optional[date],
        data_fim: Optional[date],
    ) -> List[List]:
        """Gera dados de despesas da loja."""
        despesas = db.query(DespesaLoja).all()

        rows = []
        for despesa in despesas:
            referencia = ExportacaoService._resolve_despesa_loja_date(despesa)
            if data_inicio and referencia and referencia < data_inicio:
                continue
            if data_fim and referencia and referencia > data_fim:
                continue

            data_str = referencia.strftime("%m/%Y") if referencia else ""

            row = [
                despesa.categoria or "",
                despesa.descricao or "",
                data_str,
                ExportacaoService._format_currency(despesa.valor or 0),
            ]
            rows.append(row)

        return rows

    @staticmethod
    def _generate_lancamentos_manuais(
        db: Session,
        data_inicio: Optional[date],
        data_fim: Optional[date],
    ) -> List[List]:
        """Gera dados de lançamentos financeiros manuais."""
        query = db.query(LancamentoFinanceiro)

        if data_inicio:
            query = query.filter(LancamentoFinanceiro.data >= data_inicio)
        if data_fim:
            query = query.filter(LancamentoFinanceiro.data <= data_fim)

        rows = []
        for lancamento in query.order_by(LancamentoFinanceiro.data.desc(), LancamentoFinanceiro.id.desc()).all():
            rows.append([
                ExportacaoService._format_date(lancamento.data),
                (lancamento.tipo or "").capitalize(),
                lancamento.categoria or "",
                lancamento.descricao or "",
                (lancamento.status or "").capitalize(),
                ExportacaoService._format_currency(lancamento.valor or 0),
            ])

        return rows

    @staticmethod
    def _generate_seguros(
        db: Session,
        data_inicio: Optional[date],
        data_fim: Optional[date],
    ) -> List[List]:
        """Gera dados de seguros (parcelas pagas)."""
        query = db.query(ParcelaSeguro).filter(ParcelaSeguro.status == "pago")

        if data_inicio:
            query = query.filter(ParcelaSeguro.vencimento >= data_inicio)
        if data_fim:
            query = query.filter(ParcelaSeguro.vencimento <= data_fim)

        parcelas = query.all()

        rows = []
        for parcela in parcelas:
            apólice = ""
            if parcela.seguro:
                apólice = parcela.seguro.numero_apolice or ""

            row = [
                apólice,
                ExportacaoService._format_date(parcela.vencimento),
                ExportacaoService._format_currency(parcela.valor or 0),
            ]
            rows.append(row)

        return rows

    @staticmethod
    def _generate_ipva(
        db: Session,
        data_inicio: Optional[date],
        data_fim: Optional[date],
    ) -> List[List]:
        """Gera dados de IPVA."""
        query = db.query(IpvaRegistro)

        if data_inicio:
            query = query.filter(IpvaRegistro.data_pagamento >= data_inicio)
        if data_fim:
            query = query.filter(IpvaRegistro.data_pagamento <= data_fim)

        registros = query.all()

        rows = []
        for registro in registros:
            veiculo_info = ""
            if registro.veiculo:
                veiculo_info = f"{registro.veiculo.placa or ''} {registro.veiculo.modelo or ''}".strip()

            row = [
                veiculo_info,
                str(registro.ano_referencia or ""),
                ExportacaoService._format_currency(registro.valor_ipva or 0),
                ExportacaoService._format_date(registro.data_pagamento),
            ]
            rows.append(row)

        return rows

    @staticmethod
    def _generate_multas(
        db: Session,
        data_inicio: Optional[date],
        data_fim: Optional[date],
    ) -> List[List]:
        """Gera dados de multas."""
        query = db.query(Multa)

        if data_inicio:
            query = query.filter(Multa.data_pagamento >= data_inicio)
        if data_fim:
            query = query.filter(Multa.data_pagamento <= data_fim)

        multas = query.all()

        rows = []
        for multa in multas:
            veiculo_info = ""
            if multa.veiculo:
                veiculo_info = f"{multa.veiculo.placa or ''} {multa.veiculo.modelo or ''}".strip()

            row = [
                veiculo_info,
                multa.tipo or "",
                multa.descricao or "",
                ExportacaoService._format_currency(multa.valor or 0),
                ExportacaoService._format_date(multa.data_pagamento),
            ]
            rows.append(row)

        return rows
